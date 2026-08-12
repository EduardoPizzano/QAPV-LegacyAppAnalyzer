"""Tests de la Fase 4 (IMPLEMENTATION_PLAN.md): nueva categoria de finding
para invocacion indirecta/tardia (Reflection/COM-CLSID, KNOWN_LIMITATIONS.md
L16/L17) y el patron PLC/Modbus-TCP en LOCAL_IO_TRIGGER (L18).

Decision de diseno seguida en toda esta fase (ver comentarios en
analyzer/extract.py): NO se crea una tabla nueva para reflection --
ARCHITECTURE_REVIEW.md seccion 7 ya advirtio contra multiplicar tablas
"algo que se observo sobre una app" sin discriminador. Se reutiliza
LocalIOFinding/io_findings con un campo `category` nuevo ("io" | "reflection"),
mismo patron que SqlFinding.category ya usa para distinguir query/
stored_procedure/oracle_package_call.

Los dos gaps que este archivo cierra estaban documentados como "GAP CONOCIDO"
en tests/test_characterization.py (TestDataTransferReflectionGap,
TestVins1ModbusGap) -- esos tests ya se actualizaron "a proposito" para
reflejar el comportamiento nuevo; este archivo agrega la cobertura extremo a
extremo (persistencia, reporte, exportaciones) siguiendo el mismo patron ya
usado en tests/test_increment3a_sql_reconstruction.py.
"""

from analyzer import db
from analyzer.extract import scan_project
from analyzer.report import render_from_db
from analyzer.export_office import build_docx, build_xlsx
from analyzer.techstack import TechStack


class TestReflectionDetectionInSource:
    """PrintReportViewer.cs (DataTransfer/VINS1 real): invoca miembros NO
    PUBLICOS de Microsoft.Reporting.WinForms.ReportViewer. Alcance deliberado
    (ver comentario largo en tests/test_characterization.py:
    TestDataTransferReflectionGap): se detectan los 2 puntos reales de
    invocacion (MethodInfo.Invoke, Activator.CreateInstance), no las 7
    llamadas al wrapper local `ExecuteFunction` que reenvian a ellos --
    seguir un wrapper local entre metodos esta fuera del alcance ya
    establecido para este extractor."""

    def test_scan_project_no_longer_skips_the_file(self, fixture_root):
        """Antes de esta fase, ni SQL_TRIGGER ni LOCAL_IO_TRIGGER disparaban
        en este archivo -- scan_project() lo saltaba con el filtro de vista
        previa sin abrirlo a fondo (ver REFLECTION_PREVIEW_HINT)."""
        sql_findings, io_findings = scan_project(fixture_root("datatransfer"))
        assert any("PrintReportViewer.cs" in f.file for f in io_findings)

    def test_methodinfo_invoke_detected_only_with_methodinfo_in_scope(self, fixture_root):
        """`current.Invoke(obj, parms)` -- un `.Invoke(` suelto -- solo cuenta
        como reflection porque `MethodInfo` aparece antes, en el mismo
        metodo (ExecuteFunction). Confirma que la evidencia esta atada al
        metodo real, no a un archivo entero."""
        _, io_findings = scan_project(fixture_root("datatransfer"))
        finding = next(f for f in io_findings if f.method == "ExecuteFunction")
        assert finding.category == "reflection"
        assert finding.operation == "MethodInfo.Invoke"
        assert finding.class_name == "PrintReportViewer"

    def test_activator_createinstance_detected_unambiguously(self, fixture_root):
        """Activator.CreateInstance no necesita ningun hint adicional --
        es reflection sin ambiguedad posible."""
        _, io_findings = scan_project(fixture_root("datatransfer"))
        finding = next(f for f in io_findings if f.method == "PrintByPriner")
        assert finding.category == "reflection"
        assert finding.operation == "Activator.CreateInstance"

    def test_bare_invoke_without_methodinfo_in_scope_is_not_reflection(self, fixture_root):
        """Contraprueba: un `.Invoke(` normal de delegado/evento, sin
        `MethodInfo` en el mismo metodo, NUNCA debe marcarse como reflection
        -- confirma que el hint de alcance evita falsos positivos triviales
        (ej. `myEvent.Invoke(this, EventArgs.Empty)` en cualquier app)."""
        import tempfile
        from pathlib import Path

        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            (root / "Class1.cs").write_text(
                "public class Class1 {\n"
                "    public event EventHandler MyEvent;\n"
                "    private void Fire() {\n"
                "        MyEvent.Invoke(this, EventArgs.Empty);\n"
                "    }\n"
                "}\n",
                encoding="utf-8",
            )
            sql_findings, io_findings = scan_project(root)
            assert sql_findings == []
            assert io_findings == [], (
                "Un .Invoke() de delegado comun, sin MethodInfo en el mismo metodo, "
                "no debe generar un finding de reflection"
            )


class TestModbusDetectionInSource:
    """VINS1/Modbus/Form1.cs (real): segunda integracion PLC/Modbus-TCP del
    portafolio (la primera, MonTemp2, solo documentada a mano). Categoria
    'io' -- es una integracion fisica mas, no invocacion indirecta."""

    def test_modbus_client_constructor_detected(self, fixture_root):
        sql_findings, io_findings = scan_project(fixture_root("vins1_modbus"))
        assert sql_findings == []
        assert len(io_findings) == 1
        f = io_findings[0]
        assert f.category == "io"
        assert f.operation == "new ModbusClient"
        assert "192.168.1.5" in f.raw


class TestEndToEndPersistsAndRenders:
    """Mismo patron ya usado en Incrementos 1/2/3A: la nueva capacidad debe
    sobrevivir persistencia SQLite + reconstruccion + render Markdown/Excel/
    Word, no solo funcionar en scan_project() aislado."""

    def _fresh_db(self, tmp_path, monkeypatch):
        monkeypatch.setattr(db, "DB_PATH", tmp_path / "fase4_test.db")
        db.init_db()

    def test_reflection_findings_persist_with_category(self, fixture_root, tmp_path, monkeypatch):
        self._fresh_db(tmp_path, monkeypatch)
        sql_findings, io_findings = scan_project(fixture_root("datatransfer"))
        app_id = db.save_analysis("Fase4ReflectionApp", "N/A", TechStack(), [], sql_findings, io_findings, [])

        data = db.get_app(app_id)
        reflection_rows = [r for r in data["io_findings"] if r["category"] == "reflection"]
        assert len(reflection_rows) == 2
        assert {r["operation"] for r in reflection_rows} == {"MethodInfo.Invoke", "Activator.CreateInstance"}

    def test_reflection_section_reaches_the_rendered_markdown_report(self, fixture_root, tmp_path, monkeypatch):
        self._fresh_db(tmp_path, monkeypatch)
        sql_findings, io_findings = scan_project(fixture_root("datatransfer"))
        app_id = db.save_analysis("Fase4ReflectionReport", "N/A", TechStack(), [], sql_findings, io_findings, [])

        report_md = render_from_db(db.get_app(app_id))
        assert "Invocacion indirecta / tardia (Reflection, COM)" in report_md
        assert "MethodInfo.Invoke" in report_md
        assert "Activator.CreateInstance" in report_md
        # No debe aparecer duplicado en la tabla de I/O comun.
        io_section = report_md.split("## Invocacion indirecta")[0]
        assert "MethodInfo.Invoke" not in io_section

    def test_reflection_section_reaches_the_excel_export(self, fixture_root):
        sql_findings, io_findings = scan_project(fixture_root("datatransfer"))
        xlsx_bytes = build_xlsx("Fase4ReflectionXlsx", TechStack(), [], sql_findings, io_findings, [])

        import io as iomod

        from openpyxl import load_workbook

        wb = load_workbook(iomod.BytesIO(xlsx_bytes))
        assert "Reflection-COM" in wb.sheetnames
        rows = [tuple(c.value for c in row) for row in wb["Reflection-COM"].iter_rows(min_row=2)]
        assert any(row[2] == "MethodInfo.Invoke" for row in rows), rows
        assert any(row[2] == "Activator.CreateInstance" for row in rows), rows

    def test_reflection_section_reaches_the_word_export(self, fixture_root):
        sql_findings, io_findings = scan_project(fixture_root("datatransfer"))
        docx_bytes = build_docx("Fase4ReflectionDocx", TechStack(), [], sql_findings, io_findings, [])

        import io as iomod

        from docx import Document

        doc = Document(iomod.BytesIO(docx_bytes))
        headings = [p.text for p in doc.paragraphs if p.style.name.startswith("Heading")]
        assert any("Reflection" in h for h in headings), headings

    def test_modbus_finding_persists_and_renders_in_plain_io_section(self, fixture_root, tmp_path, monkeypatch):
        self._fresh_db(tmp_path, monkeypatch)
        sql_findings, io_findings = scan_project(fixture_root("vins1_modbus"))
        app_id = db.save_analysis("Fase4ModbusApp", "N/A", TechStack(), [], sql_findings, io_findings, [])

        data = db.get_app(app_id)
        assert data["io_findings"][0]["category"] == "io"

        report_md = render_from_db(data)
        assert "192.168.1.5" in report_md
        io_section = report_md.split("## Invocacion indirecta")[0]
        assert "192.168.1.5" in io_section, "El hallazgo de Modbus debe aparecer en la tabla de I/O comun, no en la de reflection"


class TestCategoryColumnMigration:
    """Migracion aditiva de io_findings.category (mismo mecanismo ya usado
    para las columnas de Evidence en Fase 1) -- a diferencia de esas
    columnas (NULLABLE, documentan 'no instrumentado'), aqui el default 'io'
    SI es el valor real y correcto para cualquier fila anterior a esta fase,
    porque nunca existio otra categoria antes de que Reflection se
    instrumentara."""

    def test_legacy_io_findings_table_gets_category_column_with_io_default(self, tmp_path, monkeypatch):
        import sqlite3

        db_path = tmp_path / "pre_fase4.db"
        monkeypatch.setattr(db, "DB_PATH", db_path)

        # Crea la tabla en su forma ANTERIOR a esta fase (sin `category`),
        # simulando una instalacion que nunca corrio init_db() con este
        # cambio -- mismo patron ya usado por TestBackwardCompatibilityWithHistoricalRows
        # en tests/test_increment3a_sql_reconstruction.py.
        conn = sqlite3.connect(db_path)
        conn.execute(
            "CREATE TABLE apps (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL, "
            "source_path TEXT NOT NULL, analyzed_at TEXT NOT NULL)"
        )
        conn.execute(
            "CREATE TABLE io_findings (id INTEGER PRIMARY KEY AUTOINCREMENT, app_id INTEGER, "
            "file TEXT, class_name TEXT, method TEXT, operation TEXT, raw TEXT)"
        )
        conn.execute(
            "INSERT INTO apps (name, source_path, analyzed_at) VALUES ('PreFase4App', 'N/A', '2026-01-01T00:00:00')"
        )
        conn.execute(
            "INSERT INTO io_findings (app_id, file, class_name, method, operation, raw) "
            "VALUES (1, 'Old.cs', 'Old', 'OldMethod', 'File.ReadAllText', 'File.ReadAllText(\"x\")')"
        )
        conn.commit()
        conn.close()

        db.init_db()  # no debe lanzar excepcion, debe agregar la columna con default 'io'

        with db.get_conn() as conn:
            cols = {row["name"] for row in conn.execute("PRAGMA table_info(io_findings)")}
            assert "category" in cols
            row = conn.execute("SELECT category FROM io_findings WHERE file = 'Old.cs'").fetchone()
            assert row["category"] == "io", (
                "Una fila de io_findings guardada antes de esta fase debe heredar category='io' -- "
                "es el valor real conocido, no un placeholder de 'no instrumentado'"
            )

    def test_running_init_db_twice_on_migrated_db_is_safe(self, tmp_path, monkeypatch):
        monkeypatch.setattr(db, "DB_PATH", tmp_path / "idempotent.db")
        db.init_db()
        db.init_db()
        with db.get_conn() as conn:
            cols = [row["name"] for row in conn.execute("PRAGMA table_info(io_findings)")]
            assert cols.count("category") == 1


class TestComClsidRealEvidence:
    """AUDITORIA POST-CIERRE (2026-08-12), Objetivo 1: la auditoria de Fase 4
    encontro que REFLECTION_UNAMBIGUOUS ya cubre Marshal.GetTypeFromCLSID en
    codigo, pero no existia ningun fixture ni test contra un caso REAL --
    L17 quedaba "Mitigada" solo por inferencia de que el codigo se comparte
    con Reflection, nunca demostrado empiricamente.

    Fixture usado: tests/fixtures/reportviewer_imprimex_com/MainVM.cs, metodo
    ImprimeX real de decompiled/ReportViewer/.../MainVM.cs:1174-1214 -- es
    exactamente el fixture "MainVM.cs:1178" que el criterio de salida
    ORIGINAL de esta fase (IMPLEMENTATION_PLAN.md) menciono pero nunca
    congelo. Se confirmo (grep sobre decompiled/, documentado en el
    encabezado del fixture) que el mismo idioma exacto -- Activator.
    CreateInstance(Marshal.GetTypeFromCLSID(new Guid("00024500-..."))) --
    aparece IDENTICO en otras 6 apps reales del portafolio (AFLProdMon,
    SGI/INVENTA2-2TEST x2, DataTransfer, SafeRH, AppCortes/OTDR): no es un
    caso aislado, es el patron canonico de activacion tardia de
    Excel.Application via su CLSID conocido en todo el portafolio.

    HALLAZGO REAL DE ESTA INVESTIGACION (reportado, NO corregido -- fuera
    del alcance autorizado para este objetivo): en las 7 ocurrencias reales
    confirmadas, `Activator.CreateInstance(` aparece textualmente ANTES que
    `Marshal.GetTypeFromCLSID(` en la misma linea. Como REFLECTION_UNAMBIGUOUS
    usa alternancia de regex y re.search() devuelve la coincidencia mas a la
    IZQUIERDA en la cadena (no la primera alternativa del patron), el
    `operation` resultante SIEMPRE es "Activator.CreateInstance", nunca
    "Marshal.GetTypeFromCLSID" -- para este idioma real, tal como esta escrito
    hoy en el portafolio. El `category="reflection"` SI se produce
    correctamente (el proposito de riesgo de L17 se cumple), pero la
    evidencia especifica de COM/CLSID solo queda visible dentro de `raw`,
    nunca como `operation`. Este test fija el comportamiento REAL observado,
    no el que se hubiera asumido sin esta evidencia."""

    def test_com_activation_detected_as_reflection_with_real_evidence(self, fixture_root):
        sql_findings, io_findings = scan_project(fixture_root("reportviewer_imprimex_com"))
        assert sql_findings == []

        reflection_findings = [f for f in io_findings if f.category == "reflection"]
        assert len(reflection_findings) == 1
        finding = reflection_findings[0]

        assert finding.class_name == "MainVM"
        assert finding.method == "ImprimeX"
        # Comportamiento real confirmado (ver docstring de la clase): el
        # operation label es "Activator.CreateInstance", NO
        # "Marshal.GetTypeFromCLSID", porque ese texto aparece primero en la
        # linea real -- no es un supuesto, es lo que el extractor produce hoy.
        assert finding.operation == "Activator.CreateInstance"
        # La evidencia de COM/CLSID especificamente SI esta presente, en raw.
        assert "Marshal.GetTypeFromCLSID" in finding.raw
        assert "00024500-0000-0000-C000-000000000046" in finding.raw  # CLSID real de Excel.Application

    def test_com_finding_persists_with_category_reflection(self, fixture_root, tmp_path, monkeypatch):
        monkeypatch.setattr(db, "DB_PATH", tmp_path / "com_clsid_test.db")
        db.init_db()
        sql_findings, io_findings = scan_project(fixture_root("reportviewer_imprimex_com"))
        app_id = db.save_analysis("Fase4ComClsidApp", "N/A", TechStack(), [], sql_findings, io_findings, [])

        data = db.get_app(app_id)
        reflection_rows = [r for r in data["io_findings"] if r["category"] == "reflection"]
        assert len(reflection_rows) == 1
        assert "Marshal.GetTypeFromCLSID" in reflection_rows[0]["raw"]

    def test_com_finding_reaches_the_rendered_markdown_report(self, fixture_root, tmp_path, monkeypatch):
        monkeypatch.setattr(db, "DB_PATH", tmp_path / "com_clsid_report_test.db")
        db.init_db()
        sql_findings, io_findings = scan_project(fixture_root("reportviewer_imprimex_com"))
        app_id = db.save_analysis("Fase4ComClsidReport", "N/A", TechStack(), [], sql_findings, io_findings, [])

        report_md = render_from_db(db.get_app(app_id))
        assert "Invocacion indirecta / tardia (Reflection, COM)" in report_md
        assert "Marshal.GetTypeFromCLSID" in report_md
        io_section = report_md.split("## Invocacion indirecta")[0]
        assert "Marshal.GetTypeFromCLSID" not in io_section, (
            "El hallazgo COM/CLSID no debe aparecer duplicado en la tabla de I/O comun"
        )


class TestDiagramReflectionCategoryLabel:
    """AUDITORIA POST-CIERRE (2026-08-12), Objetivo 2: diagram.py ya contiene
    la rama `if f.category == "reflection": return "Reflection / COM"` en
    _io_category(), pero ningun test la ejercitaba (confirmado por grep --
    cero referencias a `diagram`/`build_dataflow` en toda la suite antes de
    este archivo). Test minimo de cobertura de la rama ya existente, sin
    fixture de decompilado (build_dataflow_diagram es una transformacion
    pura sobre hallazgos ya extraidos, igual que documenta su propio
    docstring de modulo -- no necesita decompilar nada)."""

    def test_reflection_finding_gets_its_own_diagram_node_label(self):
        from analyzer.diagram import build_dataflow_diagram
        from analyzer.extract import LocalIOFinding

        finding = LocalIOFinding(
            file="MainVM.cs",
            class_name="MainVM",
            method="ImprimeX",
            operation="Activator.CreateInstance",
            raw="Activator.CreateInstance(Marshal.GetTypeFromCLSID(...))",
            category="reflection",
        )
        diagram_text = build_dataflow_diagram([], [finding])

        assert diagram_text is not None
        assert "Reflection / COM" in diagram_text
        # No debe caer en el catch-all "Otro I/O" -- confirma que la rama de
        # category se evalua ANTES que los prefijos de IO_CATEGORY_PATTERNS.
        assert "Otro I/O" not in diagram_text

    def test_plain_io_finding_is_not_mislabeled_as_reflection(self):
        """Contraprueba: un hallazgo de I/O comun (category='io') nunca debe
        etiquetarse como 'Reflection / COM' en el diagrama, ni siquiera si su
        `operation` contuviera coincidentalmente una palabra parecida."""
        from analyzer.diagram import build_dataflow_diagram
        from analyzer.extract import LocalIOFinding

        finding = LocalIOFinding(
            file="Form1.cs", class_name="Form1", method="button1_Click",
            operation="new ModbusClient", raw="new ModbusClient(\"192.168.1.5\")", category="io",
        )
        diagram_text = build_dataflow_diagram([], [finding])

        assert diagram_text is not None
        assert "Reflection / COM" not in diagram_text
        assert "PLC / Modbus" in diagram_text


class TestMixedIoAndReflectionCategoriesCoexist:
    """AUDITORIA POST-CIERRE (2026-08-12), Objetivo 4: gap no bloqueante
    identificado por la auditoria -- ningun test previo cubria una MISMA app
    con category='io' y category='reflection' simultaneamente. Reutiliza los
    dos fixtures ya existentes (datatransfer, vins1_modbus) combinando sus
    resultados en un unico analisis, sin crear ninguna infraestructura de
    testing nueva."""

    def test_both_categories_persist_and_render_without_mixing_or_duplication(
        self, fixture_root, tmp_path, monkeypatch
    ):
        monkeypatch.setattr(db, "DB_PATH", tmp_path / "mixed_categories_test.db")
        db.init_db()

        _, reflection_io = scan_project(fixture_root("datatransfer"))
        _, plain_io = scan_project(fixture_root("vins1_modbus"))
        combined_io_findings = [f for f in reflection_io if f.category == "reflection"] + plain_io

        assert any(f.category == "reflection" for f in combined_io_findings)
        assert any(f.category == "io" for f in combined_io_findings)

        app_id = db.save_analysis("Fase4MixedCategoriesApp", "N/A", TechStack(), [], [], combined_io_findings, [])

        data = db.get_app(app_id)
        db_categories = [r["category"] for r in data["io_findings"]]
        assert db_categories.count("reflection") == 2  # MethodInfo.Invoke + Activator.CreateInstance
        assert db_categories.count("io") == 1  # new ModbusClient
        assert len(db_categories) == 3, "No debe haber duplicacion de hallazgos al combinar ambas categorias"

        report_md = render_from_db(data)
        reflection_section = report_md.split("## Invocacion indirecta")[1]
        io_section = report_md.split("## Invocacion indirecta")[0]

        assert "MethodInfo.Invoke" in reflection_section
        assert "Activator.CreateInstance" in reflection_section
        assert "192.168.1.5" in io_section
        # Ninguna mezcla cruzada entre secciones.
        assert "192.168.1.5" not in reflection_section
        assert "MethodInfo.Invoke" not in io_section
        assert "Activator.CreateInstance" not in io_section
