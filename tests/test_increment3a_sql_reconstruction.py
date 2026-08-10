"""Tests del Incremento Funcional 3A: resolucion de SQL estatico y
multilinea construido via variable/StringBuilder DENTRO DEL MISMO METODO
(literal simple, multilinea, verbatim, concatenacion '+', StringBuilder.
Append/AppendLine/ToString). Alcance explicito de VALIDATION_STRATEGY.md:
nada de analisis entre metodos, Reflection, Stored Procedures, ejecucion
simbolica -- los casos con ramificacion (if/else/for/while/switch/ternario)
deben QUEDAR sin resolver a proposito, nunca fabricar un valor.

TestSgiStringBuilderGap y TestDataTransferStringBuilderGap en
test_characterization.py ya cubren, respectivamente, el caso StringBuilder
CON ramificacion (sigue sin resolver) y SIN ramificacion (ya resuelto) --
este archivo agrega los casos nuevos (concatenacion simple, verbatim
multilinea, ternario) y valida la Evidence/Confidence/persistencia/reporte
end-to-end, mismo patron ya usado en los Incrementos 1 y 2.
"""

from datetime import datetime

import pytest

from analyzer import db
from analyzer.__version__ import ANALYZER_VERSION
from analyzer.confidence import CONFIDENCE_TABLE
from analyzer.export_office import build_docx, build_xlsx
from analyzer.extract import scan_project
from analyzer.report import _group_by_method, _rows_for_method, render_from_db
from analyzer.techstack import TechStack


class TestConcatenationResolvesWithRealEvidence:
    """AFL.Dashboard/Class1.cs:UpdateJobLinea -- `new SqlCommand(cmdText,
    sqlConnection)` (dos argumentos) es el patron dominante real del
    portafolio; antes de este incremento ningun regex de deteccion de
    variable lo reconocia, asi que nunca se intentaba resolver."""

    def test_concatenation_reconstructs_with_placeholders_for_dynamic_parts(self, fixture_root):
        sql_findings, _ = scan_project(fixture_root("concat_case"))
        resolved = [f for f in sql_findings if f.target == "LCJob"]
        assert len(resolved) == 1
        f = resolved[0]
        assert f.resolved == "Update LCJob set Linea='{linea}',EntregadoA='{entregadoA}'  where ID={idJob}"

    def test_evidence_reflects_partial_reconstruction_not_full_literal(self, fixture_root):
        """Quedan segmentos dinamicos (variables C# reales) -- confidence
        debe ser PARTIAL_RECONSTRUCTION (80), NUNCA HARDCODED_METHOD_LITERAL
        (90, reservado para cuando TODO el valor es literal conocido)."""
        sql_findings, _ = scan_project(fixture_root("concat_case"))
        f = next(x for x in sql_findings if x.target == "LCJob")
        ev = f.evidence
        assert ev.extractor == "PARTIAL_RECONSTRUCTION"
        assert ev.confidence == CONFIDENCE_TABLE["PARTIAL_RECONSTRUCTION"] == 80
        assert ev.pattern == "STRING_VAR_ASSIGN"
        assert ev.line_number == 21  # linea real de `string cmdText = ...` en el fixture
        assert "cmdText" in ev.snippet
        assert ev.analyzer_version == ANALYZER_VERSION
        datetime.fromisoformat(ev.created_at)


class TestFullyLiteralReconstructionGetsHighestConfidence:
    """AFL.Dashboard/Class1.cs:GrabaOperacion -- un INSERT ya parametrizado
    con @params, sin ninguna concatenacion. Reconstruido 100% literal debe
    valer igual que un `var = "...";` simple (HARDCODED_METHOD_LITERAL)."""

    def test_pure_concatenation_of_literals_is_full_confidence(self, fixture_root):
        """Caso limite: si CADA segmento de la concatenacion es literal (sin
        ninguna variable de por medio), el resultado es HARDCODED_METHOD_LITERAL,
        no PARTIAL_RECONSTRUCTION -- se sabe el 100% del contenido."""
        sql_findings, _ = scan_project(fixture_root("verbatim_multiline_case"))
        f = next(x for x in sql_findings if x.target == "LCJob")
        assert f.evidence.extractor == "HARDCODED_METHOD_LITERAL"
        assert f.evidence.confidence == CONFIDENCE_TABLE["HARDCODED_METHOD_LITERAL"] == 90


class TestVerbatimMultilineStringResolves:
    def test_verbatim_multiline_literal_reconstructs_with_real_newlines(self, fixture_root):
        sql_findings, _ = scan_project(fixture_root("verbatim_multiline_case"))
        f = next(x for x in sql_findings if x.target == "LCJob")
        assert "UPDATE LCJob" in f.resolved
        assert "SET IDEstatus = 4" in f.resolved
        assert "\n" in f.resolved, "El salto de linea real dentro del verbatim string debe preservarse"


class TestBranchingNeverGetsResolvedNeverFabricated:
    """Alcance explicito: nada de ejecucion simbolica. Un ternario (o un
    StringBuilder con if/else, ya cubierto en TestSgiStringBuilderGap de
    test_characterization.py) implica que el valor final depende de una
    condicion no evaluada -- debe quedar sin resolver, nunca a medias."""

    def test_ternary_assigned_variable_stays_unresolved(self, fixture_root):
        sql_findings, _ = scan_project(fixture_root("ternary_branch_case"))
        command_findings = [f for f in sql_findings if f.kind == "SqlCommand"]
        assert command_findings
        for f in command_findings:
            assert f.resolved is None
            assert f.target is None
            assert f.evidence.extractor == "UNKNOWN"  # default de Fase 1, nunca inventado

    def test_report_shows_generic_message_not_a_fabricated_partial_value(self, fixture_root):
        sql_findings, _ = scan_project(fixture_root("ternary_branch_case"))
        groups = _group_by_method(sql_findings)
        group = groups[("Class1", "BuscaCantidad")]
        rows = list(_rows_for_method(group))
        assert len(rows) == 1
        assert rows[0][0] == "(conexion detectada, query no resuelta automaticamente — revisar manualmente)"


class TestEndToEndPersistsAndRenders:
    @pytest.fixture
    def fresh_db(self, tmp_path, monkeypatch):
        monkeypatch.setattr(db, "DB_PATH", tmp_path / "increment3a_test.db")
        db.init_db()

    def test_reconstructed_sql_persists_over_existing_columns_no_schema_change(self, fixture_root, fresh_db):
        """sql_findings ya tenia las columnas de Evidence desde el Incremento
        1 (nunca se les escribia) -- este incremento es el primero que las
        llena para sql_findings, sin migrar el esquema otra vez."""
        sql_findings, _ = scan_project(fixture_root("concat_case"))
        app_id = db.save_analysis("Increment3aConcatCase", "N/A", TechStack(), [], sql_findings, [], [])

        data = db.get_app(app_id)
        row = next(r for r in data["sql_findings"] if r["target"] == "LCJob")
        assert row["extractor"] == "PARTIAL_RECONSTRUCTION"
        assert row["pattern"] == "STRING_VAR_ASSIGN"
        assert row["confidence"] == 80
        assert row["line_number"] == 21
        assert row["resolved"] == "Update LCJob set Linea='{linea}',EntregadoA='{entregadoA}'  where ID={idJob}"

    def test_evidence_reaches_the_rendered_markdown_report(self, fixture_root, fresh_db):
        sql_findings, _ = scan_project(fixture_root("concat_case"))
        app_id = db.save_analysis("Increment3aConcatCaseReport", "N/A", TechStack(), [], sql_findings, [], [])

        report_md = render_from_db(db.get_app(app_id))

        assert "Update LCJob set Linea=" in report_md
        assert "PARTIAL_RECONSTRUCTION" in report_md
        assert "80%" in report_md
        assert "linea 21" in report_md

    def test_evidence_reaches_the_real_flask_route(self, fixture_root, fresh_db, monkeypatch):
        import app as flaskapp

        monkeypatch.setattr(flaskapp.db, "DB_PATH", db.DB_PATH)

        sql_findings, _ = scan_project(fixture_root("concat_case"))
        app_id = db.save_analysis("Increment3aConcatCaseLive", "N/A", TechStack(), [], sql_findings, [], [])

        client = flaskapp.app.test_client()
        resp = client.get(f"/apps/{app_id}")

        assert resp.status_code == 200
        html = resp.get_data(as_text=True)
        assert "Update LCJob set Linea=" in html
        assert "PARTIAL_RECONSTRUCTION" in html

    def test_evidence_reaches_the_excel_export(self, fixture_root):
        sql_findings, _ = scan_project(fixture_root("concat_case"))
        xlsx_bytes = build_xlsx("Increment3aXlsx", TechStack(), [], sql_findings, [], [])

        import io

        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["Funciones-SQL-SP"]
        rows = [tuple(cell.value for cell in row) for row in ws.iter_rows(min_row=2)]
        assert any(
            row[4] == "LCJob" and row[8] == "PARTIAL_RECONSTRUCTION" and row[10] == 80
            for row in rows
        ), rows

    def test_evidence_reaches_the_word_export(self, fixture_root):
        sql_findings, _ = scan_project(fixture_root("concat_case"))
        docx_bytes = build_docx("Increment3aDocx", TechStack(), [], sql_findings, [], [])

        import io

        from docx import Document

        doc = Document(io.BytesIO(docx_bytes))
        table = next(t for t in doc.tables if t.rows[0].cells[0].text == "Clase")
        body_rows = [[c.text for c in row.cells] for row in table.rows[1:]]
        assert any("LCJob" in row and "PARTIAL_RECONSTRUCTION" in row for row in body_rows), body_rows


class TestBackwardCompatibilityWithHistoricalRows:
    """Una fila de sql_findings guardada ANTES de este incremento no tiene
    estas columnas pobladas -- debe seguir renderizando el default honesto
    de Fase 1, nunca romper el reporte."""

    @pytest.fixture
    def fresh_db(self, tmp_path, monkeypatch):
        monkeypatch.setattr(db, "DB_PATH", tmp_path / "increment3a_compat_test.db")
        db.init_db()

    def test_sql_finding_without_evidence_columns_renders_default(self, fresh_db):
        with db.get_conn() as conn:
            conn.execute(
                "INSERT INTO apps (name, source_path, analyzed_at) VALUES (?, ?, ?)",
                ("PreExistingSqlFindingApp", "N/A", "2026-01-01T00:00:00"),
            )
            app_id = conn.execute(
                "SELECT id FROM apps WHERE name = ?", ("PreExistingSqlFindingApp",)
            ).fetchone()["id"]
            conn.execute(
                "INSERT INTO sql_findings (app_id, file, class_name, method, kind, category, target, "
                "is_stored_procedure, raw, resolved) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (app_id, "Old.cs", "Old", "OldMethod", "CommandText", "query", "OldTable", 0,
                 'sqlCommand.CommandText = "SELECT * FROM OldTable";',
                 "SELECT * FROM OldTable"),
            )

        report_md = render_from_db(db.get_app(app_id))
        assert "UNKNOWN" in report_md
        assert "20%" in report_md
