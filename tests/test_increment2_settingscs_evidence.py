"""Tests del Incremento Funcional 2: instrumentacion completa de Settings.cs
(DefaultSettingValue) con el mismo patron ya validado para app.config en el
Incremento Funcional 1 -- Evidence real en el momento de la extraccion,
persistencia sobre las columnas existentes (sin migrar esquema otra vez),
reporte (Markdown/HTML) y exportes (Excel/Word) consistentes, y compatibilidad
con analisis historicos que no tienen estas columnas pobladas.

Nota de diseno (documentada aqui, no solo en el commit): el pedido original
menciona `extractor="SETTINGS_CLASS_CONNECTION"` y `pattern="Settings.Designer.cs"`.
Ninguno de los dos se usa tal cual:
- `SETTINGS_CLASS_CONNECTION` no existe en `analyzer.confidence.CONFIDENCE_TABLE`
  (Fase 1, ya aprobada) -- usarlo haria que `resolve_confidence()` cayera al
  piso UNKNOWN=20 en vez del 95 que el propio catalogo ya asigna a este
  mecanismo exacto bajo el nombre `SETTINGS_DEFAULT_VALUE` ("el mecanismo
  dominante de descubrimiento de conexiones en este portafolio", segun el
  comentario ya existente en confidence.py). Usar el nombre nuevo hubiera
  sido asignar confidence a mano por la puerta de atras -- exactamente lo que
  la regla de diseno prohibe.
- ilspycmd emite el archivo como `Settings.cs` en este portafolio (confirmado
  en los fixtures `happy_path`/`dedup_case`, y documentado en el comentario
  del fixture `happy_path` citando un ejemplo real decompilado), nunca
  `Settings.Designer.cs` -- ese nombre ya se captura en `source_file` de todas
  formas, asi que `pattern` describe la regla que disparo (`DefaultSettingValue`,
  mismo estilo que `connectionStrings/add` del Incremento 1) en vez de repetir
  el nombre del archivo en otro campo.
"""

from datetime import datetime

import pytest

from analyzer import db
from analyzer.__version__ import ANALYZER_VERSION
from analyzer.confidence import CONFIDENCE_TABLE
from analyzer.export_office import build_docx, build_xlsx
from analyzer.extract import find_settings
from analyzer.report import render_from_db
from analyzer.techstack import TechStack


class TestSettingsCsExtractionCarriesRealEvidence:
    def test_happy_path_connection_has_real_evidence(self, fixture_root):
        settings = find_settings(fixture_root("happy_path"))
        conn = next(s for s in settings if s.name == "CX")
        ev = conn.evidence

        assert ev.extractor == "SETTINGS_DEFAULT_VALUE"
        assert ev.pattern == "DefaultSettingValue"
        assert ev.confidence == CONFIDENCE_TABLE["SETTINGS_DEFAULT_VALUE"] == 95
        assert ev.source_file.endswith("Settings.cs")
        assert ev.line_number == 25  # linea real del atributo en el fixture
        assert "DefaultSettingValue" in ev.snippet
        assert ev.analyzer_version == ANALYZER_VERSION
        datetime.fromisoformat(ev.created_at)  # ISO real, no el default None

    def test_dedup_case_surviving_entry_keeps_real_evidence(self, fixture_root):
        """La deduplicacion por valor (Fase 0) ocurre DESPUES de construir el
        Evidence -- confirma que la entrada que sobrevive (la de Settings.cs)
        no perdio su evidencia en el camino."""
        settings = find_settings(fixture_root("dedup_case"))
        assert len(settings) == 1
        assert settings[0].evidence.extractor == "SETTINGS_DEFAULT_VALUE"


class TestSettingsCsEvidencePersistsAndRenders:
    @pytest.fixture
    def fresh_db(self, tmp_path, monkeypatch):
        monkeypatch.setattr(db, "DB_PATH", tmp_path / "increment2_test.db")
        db.init_db()

    def test_persists_over_existing_columns_no_schema_change(self, fixture_root, fresh_db):
        """No se crea ninguna columna/tabla nueva -- mismas columnas que ya
        uso el Incremento 1 para app.config."""
        settings = find_settings(fixture_root("happy_path"))
        app_id = db.save_analysis("Increment2HappyPath", "N/A", TechStack(), settings, [], [], [])

        data = db.get_app(app_id)
        row = next(s for s in data["settings"] if s["name"] == "CX")

        assert row["extractor"] == "SETTINGS_DEFAULT_VALUE"
        assert row["pattern"] == "DefaultSettingValue"
        assert row["confidence"] == 95
        assert row["line_number"] == 25
        assert row["analyzer_version"] == ANALYZER_VERSION
        assert row["created_at"] is not None

    def test_evidence_reaches_the_rendered_markdown_report(self, fixture_root, fresh_db):
        settings = find_settings(fixture_root("happy_path"))
        app_id = db.save_analysis("Increment2HappyPathReport", "N/A", TechStack(), settings, [], [], [])

        report_md = render_from_db(db.get_app(app_id))

        assert "SETTINGS_DEFAULT_VALUE" in report_md
        assert "95%" in report_md
        assert "linea 25" in report_md

    def test_evidence_reaches_the_real_flask_route(self, fixture_root, fresh_db, monkeypatch):
        """Mismo mecanismo de validacion en vivo usado en el Incremento 1:
        se ejercita la ruta real de app.py (no solo render_from_db en
        aislamiento) via test_client(), sin tocar la BD de produccion."""
        import app as flaskapp

        monkeypatch.setattr(flaskapp.db, "DB_PATH", db.DB_PATH)

        settings = find_settings(fixture_root("happy_path"))
        app_id = db.save_analysis("Increment2HappyPathLive", "N/A", TechStack(), settings, [], [], [])

        client = flaskapp.app.test_client()
        resp = client.get(f"/apps/{app_id}")

        assert resp.status_code == 200
        html = resp.get_data(as_text=True)
        assert "SETTINGS_DEFAULT_VALUE" in html
        assert "95%" in html

    def test_evidence_reaches_the_excel_export(self, fixture_root):
        settings = find_settings(fixture_root("happy_path"))
        xlsx_bytes = build_xlsx("Increment2HappyPathXlsx", TechStack(), settings, [], [], [])

        import io

        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["Conexiones y config"]
        rows = [tuple(cell.value for cell in row) for row in ws.iter_rows(min_row=2)]
        assert any(row[4] == "SETTINGS_DEFAULT_VALUE" and row[6] == 95 for row in rows), rows

    def test_evidence_reaches_the_word_export(self, fixture_root):
        settings = find_settings(fixture_root("happy_path"))
        docx_bytes = build_docx("Increment2HappyPathDocx", TechStack(), settings, [], [], [])

        import io

        from docx import Document

        doc = Document(io.BytesIO(docx_bytes))
        table = next(t for t in doc.tables if t.rows[0].cells[0].text == "Setting")
        body_rows = [[c.text for c in row.cells] for row in table.rows[1:]]
        assert any("SETTINGS_DEFAULT_VALUE" in row and "95" in row for row in body_rows), body_rows


class TestBackwardCompatibilityWithHistoricalRows:
    """Una fila guardada ANTES de este incremento (o del Incremento 1) no
    tiene estas columnas pobladas -- debe seguir funcionando, mostrando el
    default honesto de Fase 1, nunca inventando un valor ni rompiendo el
    reporte."""

    @pytest.fixture
    def fresh_db(self, tmp_path, monkeypatch):
        monkeypatch.setattr(db, "DB_PATH", tmp_path / "increment2_compat_test.db")
        db.init_db()

    def test_setting_without_evidence_columns_renders_default(self, fresh_db):
        """Un SettingEntry() nuevo con Evidence por defecto SI persiste
        ("UNKNOWN"/20, ver Fase 1) -- eso no es el caso a proteger aqui. El
        caso real de "analisis historico" es una fila insertada por codigo
        ANTERIOR a este incremento, que nunca escribio estas columnas -- NULL
        real en la BD. Se inserta directo por SQL (sin pasar por
        save_analysis, que ya escribe estas columnas siempre) para replicar
        exactamente esa fila vieja."""
        with db.get_conn() as conn:
            conn.execute(
                "INSERT INTO apps (name, source_path, analyzed_at) VALUES (?, ?, ?)",
                ("PreExistingEvidenceApp", "N/A", "2026-01-01T00:00:00"),
            )
            app_id = conn.execute("SELECT id FROM apps WHERE name = ?", ("PreExistingEvidenceApp",)).fetchone()["id"]
            conn.execute(
                "INSERT INTO settings (app_id, name, default_value, is_connection_string, category, source_file) "
                "VALUES (?, ?, ?, ?, ?, ?)",
                (app_id, "CX", "Server=OLD-SERVER; Database=Old;", 1, "sql_or_oracle", "Settings.cs"),
            )

        data = db.get_app(app_id)
        row = next(s for s in data["settings"] if s["name"] == "CX")
        assert row["extractor"] is None  # NULL real en la BD, no inventado

        report_md = render_from_db(data)
        assert "UNKNOWN" in report_md
        assert "20%" in report_md
